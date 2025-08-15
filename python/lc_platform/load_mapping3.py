import os
import re
from openpyxl import Workbook


def load_mapping(mapping_file):
    """加载表名映射字典：原始名（大写） -> 新表名（小写）"""
    mapping = {}
    with open(mapping_file, "r", encoding="utf-8") as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) == 2:
                mapping[parts[0].upper()] = parts[1].lower()
    return mapping


def process_folder(folder_path, mapping_file, output_xlsx, output_log_txt):
    mapping = load_mapping(mapping_file)

    # 正则：匹配 SQL 中用于替换的表名（出现在 insert into / from / join / update 后）
    sql_pattern = re.compile(
        r"\b(insert\s+into|update|from|join|delete\s+from)\s+(OTC142\.|OTC\.)([A-Za-z0-9_]+)",
        re.IGNORECASE,
    )
    # 正则：匹配 .NEXTVAL 序列
    seq_pattern = re.compile(
        r"(OTC142\.|OTC\.)([A-Za-z0-9_]+\.)*[A-Za-z0-9_]+\.NEXTVAL", re.IGNORECASE
    )

    # 日志用
    all_replaced = set()
    all_sequences = set()
    all_no_replace = set()
    log_lines = []

    for root, _, files in os.walk(folder_path):
        for fname in files:
            if not fname.endswith(".xml"):
                continue

            full_path = os.path.join(root, fname)
            with open(full_path, "r", encoding="utf-8") as f:
                content = f.read()

            # 1️⃣ 替换 .NEXTVAL 为 __PLACEHOLDER__，防止它被误替换
            seq_placeholders = {}

            def replace_seq(m):
                placeholder = f"__SEQ_PLACEHOLDER_{len(seq_placeholders)}__"
                seq_placeholders[placeholder] = m.group(0)
                all_sequences.add(m.group(0).upper())
                return placeholder

            content_masked = seq_pattern.sub(replace_seq, content)

            # 2️⃣ 替换 SQL 里的表名（只在特定语句中）
            replaced_info = []
            no_replace_info = []

            def replace_table(m):
                keyword = m.group(1)
                prefix = m.group(2)
                table = m.group(3)
                table_upper = table.upper()
                if table_upper in mapping:
                    new_table = mapping[table_upper]
                    replaced_info.append(
                        (prefix.upper(), table_upper, "lcinfo_source.", new_table)
                    )
                    return f"{keyword} lcinfo_source.{new_table}"
                else:
                    no_replace_info.append((prefix.upper(), table_upper))
                    return m.group(0)

            new_content = sql_pattern.sub(replace_table, content_masked)

            # 3️⃣ 恢复占位的 .NEXTVAL 序列
            for ph, original in seq_placeholders.items():
                new_content = new_content.replace(ph, original)

            # 4️⃣ 写入文件
            with open(full_path, "w", encoding="utf-8") as f:
                f.write(new_content)

            # 记录日志
            all_replaced.update(replaced_info)
            all_no_replace.update(no_replace_info)

            log_lines.append(f"=== 文件: {fname} ===")
            if replaced_info:
                log_lines.append("替换成功表名:")
                for r in replaced_info:
                    log_lines.append(f"{r[0]} {r[1]} {r[2]} {r[3]}")
            else:
                log_lines.append("无替换成功表名")

            if seq_placeholders:
                log_lines.append("\n序列表名:")
                for s in seq_placeholders.values():
                    log_lines.append(s)
            else:
                log_lines.append("\n无序列表名")

            if no_replace_info:
                log_lines.append("\n未替换表名:")
                for r in no_replace_info:
                    log_lines.append(f"{r[0]} {r[1]}")
            else:
                log_lines.append("\n无未替换表名")
            log_lines.append("\n")

    # 写日志
    with open(output_log_txt, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))

    # 写 Excel
    wb = Workbook()

    # Sheet1: 替换成功
    ws1 = wb.active
    ws1.title = "替换成功表名"
    ws1.append(["原前缀", "原表名", "新前缀", "新表名"])
    for item in sorted(all_replaced):
        ws1.append(item)

    # Sheet2: 序列
    ws2 = wb.create_sheet("序列表名")
    ws2.append(["序列名 (含.NEXTVAL)"])
    for item in sorted(all_sequences):
        ws2.append([item])

    # Sheet3: 未替换
    ws3 = wb.create_sheet("未替换表名")
    ws3.append(["前缀", "表名"])
    for item in sorted(all_no_replace):
        ws3.append(item)

    wb.save(output_xlsx)


# === 执行入口 ===
if __name__ == "__main__":
    folder_path = "/Users/mac/dev/code/py-tools/parse_xml_get_table_name/xml"
    mapping_file = (
        "/Users/mac/dev/code/py-tools/parse_xml_get_table_name/xml/table_mapping.txt"
    )
    output_xlsx = "/Users/mac/dev/code/py-tools/parse_xml_get_table_name/output/replace_summary.xlsx"
    output_log_txt = (
        "/Users/mac/dev/code/py-tools/parse_xml_get_table_name/output/replace_log.txt"
    )

    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    process_folder(folder_path, mapping_file, output_xlsx, output_log_txt)
