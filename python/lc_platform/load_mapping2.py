import os
import re
from openpyxl import Workbook


def load_mapping(mapping_file):
    mapping = {}
    with open(mapping_file, "r", encoding="utf-8") as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) == 2:
                mapping[parts[0].upper()] = parts[1].lower()
    return mapping


def process_folder(folder_path, mapping_file, output_xlsx, output_log_txt):
    mapping = load_mapping(mapping_file)

    seq_pattern = re.compile(
        r"(OTC142\.|OTC\.)((?:[A-Za-z0-9_]+\.)*[A-Za-z0-9_]+)\.NEXTVAL(?![A-Za-z0-9_])",
        re.IGNORECASE,
    )

    # 全局集合，后续写Excel用
    all_replaced = (
        set()
    )  # (old_prefix.upper(), old_table.upper(), new_prefix.lower(), new_table.lower())
    all_sequences = set()  # 全序列名，统一大写
    all_no_replaced = set()  # (prefix.upper(), table.upper())

    log_lines = []

    for root, _, files in os.walk(folder_path):
        for fname in files:
            if not fname.endswith(".xml"):
                continue
            full_path = os.path.join(root, fname)
            with open(full_path, "r", encoding="utf-8") as f:
                content = f.read()

            replaced_info = []
            seq_tables = []
            no_replace_tables = []

            # 找序列，避免误替换
            seq_spans = []
            for m in seq_pattern.finditer(content):
                full_seq = m.group(0)
                seq_tables.append(full_seq)
                seq_spans.append((m.start(), m.end()))

            # 用空格占位序列部分
            content_masked = []
            last_idx = 0
            for start, end in seq_spans:
                content_masked.append(content[last_idx:start])
                content_masked.append(" " * (end - start))
                last_idx = end
            content_masked.append(content[last_idx:])
            content_masked = "".join(content_masked)

            # def replacer(m):
            #     prefix = m.group(1)
            #     table = m.group(2)
            #     table_upper = table.upper()

            #     if table_upper in mapping:
            #         new_table = mapping[table_upper]
            #         replaced_info.append((prefix.upper(), table_upper, 'lcinfo_source.', new_table))
            #         return 'lcinfo_source.' + new_table
            #     else:
            #         no_replace_tables.append((prefix.upper(), table_upper))
            #         return m.group(0)

            # new_content = table_name_pattern.sub(replacer, content_masked)
            # 只在 SQL 关键字后面出现的表名做替换（避免误替换标签属性）
            def sql_replacer(m):
                keyword = m.group(1)
                prefix = m.group(2)
                table = m.group(3)
                table_upper = table.upper()

                if table_upper in mapping:
                    new_table = mapping[table_upper]
                    replaced_info.append(
                        (prefix.upper(), table_upper, "lcinfo_source.", new_table)
                    )
                    return f"{keyword} lcinfo_source.{new_table}"
                else:
                    no_replace_tables.append((prefix.upper(), table_upper))
                    return m.group(0)

            sql_pattern = re.compile(
                r"\b(insert\s+into|update|from|join|delete\s+from)\s+(OTC142\.|OTC\.)([A-Za-z0-9_]+)",
                re.IGNORECASE,
            )

            new_content = sql_pattern.sub(sql_replacer, content_masked)

            # 还原序列
            def restore_sequences(text, original, spans):
                result = []
                last_idx = 0
                for start, end in spans:
                    result.append(text[last_idx:start])
                    result.append(original[start:end])
                    last_idx = end
                result.append(text[last_idx:])
                return "".join(result)

            new_content = restore_sequences(new_content, content, seq_spans)

            with open(full_path, "w", encoding="utf-8") as f:
                f.write(new_content)

            # 汇总全局
            for r in replaced_info:
                # 去大小写统一存储
                all_replaced.add(
                    (r[0].upper(), r[1].upper(), r[2].lower(), r[3].lower())
                )
            for s in seq_tables:
                all_sequences.add(s.upper())
            for nr in no_replace_tables:
                all_no_replaced.add((nr[0].upper(), nr[1].upper()))

            # 写日志，按文件
            log_lines.append(f"=== 文件: {fname} ===")
            if replaced_info:
                log_lines.append("替换成功表名:")
                # 去重处理
                seen = set()
                for r in replaced_info:
                    if r not in seen:
                        seen.add(r)
                        log_lines.append(f"{r[0]} {r[1]} {r[2]} {r[3]}")
            else:
                log_lines.append("无替换成功表名")

            if seq_tables:
                log_lines.append("\n序列表名（含.NEXTVAL，未替换）:")
                # 去重，保持顺序
                seen_seq = set()
                seq_unique = []
                for s in seq_tables:
                    if s not in seen_seq:
                        seen_seq.add(s)
                        seq_unique.append(s)
                log_lines.extend(seq_unique)
            else:
                log_lines.append("\n无序列表名")

            if no_replace_tables:
                log_lines.append("\n未替换表名（不含.NEXTVAL）:")
                # 去重，保持顺序
                seen_nr = set()
                nr_unique = []
                for nr in no_replace_tables:
                    if nr not in seen_nr:
                        seen_nr.add(nr)
                        nr_unique.append(f"{nr[0]} {nr[1]}")
                log_lines.extend(nr_unique)
            else:
                log_lines.append("\n无未替换表名")

            log_lines.append("\n")

    # 写日志文件
    with open(output_log_txt, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))

    # 写 Excel
    wb = Workbook()

    # 替换表名 sheet
    ws1 = wb.active
    ws1.title = "替换成功表名"
    ws1.append(["原前缀", "原表名", "新前缀", "新表名"])
    # 按字母排序，方便查看
    for item in sorted(all_replaced, key=lambda x: (x[0], x[1])):
        ws1.append(item)

    # 序列表名 sheet
    ws2 = wb.create_sheet("序列表名")
    ws2.append(["序列全名 (含.NEXTVAL)"])
    for seq in sorted(all_sequences):
        ws2.append([seq])
    # 先提取序列对应的普通表名（去掉.NEXTVAL）
    seq_tables_without_nextval = set()
    for seq_fullname in all_sequences:
        if seq_fullname.endswith(".NEXTVAL"):
            seq_without_nextval = seq_fullname[:-7]  # 去掉 .NEXTVAL
            seq_tables_without_nextval.add(seq_without_nextval.upper())

    # 过滤未替换表名，排除序列中的表名
    filtered_no_replaced = set()
    for prefix, table in all_no_replaced:
        combined = f"{prefix}.{table}".upper()
        if combined not in seq_tables_without_nextval:
            filtered_no_replaced.add((prefix, table))
        # 未替换表名 sheet
    ws3 = wb.create_sheet("未替换表名")
    ws3.append(["前缀", "表名"])
    for nr in sorted(filtered_no_replaced, key=lambda x: (x[0], x[1])):
        ws3.append(nr)

    wb.save(output_xlsx)


if __name__ == "__main__":
    folder_path = "/Users/mac/dev/code/py-tools/parse_xml_get_table_name/xml"
    mapping_file = (
        "/Users/mac/dev/code/py-tools/parse_xml_get_table_name/xml/table_mapping.txt"
    )
    output_xlsx = "/Users/mac/dev/code/py-tools/parse_xml_get_table_name/output/replace_summary.xlsx"
    output_log_txt = (
        "/Users/mac/dev/code/py-tools/parse_xml_get_table_name/output/replace_log.txt"
    )

    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    process_folder(folder_path, mapping_file, output_xlsx, output_log_txt)
